#!/usr/bin/env python3
"""
将 cases.json 填充进测试用例 Excel 模板。

用法:
    python3 fill_template.py \
        --template assets/测试用例模版.xlsx \
        --input  测试用例文档/{模块}/{模块}-cases.json \
        --output 测试用例文档/{模块}/{模块}-核心流程测试用例-YYYYMMDD.xlsx

约束:
    - 不重新构造 xlsx, 直接复制模板, 保留样式/列宽/表头
    - 表头固定在第 1 行, 数据从第 2 行开始追加
    - steps 与 expects 数组长度必须相等, 否则报错退出
    - 测试步骤 / 预期结果列启用 wrap_text, 自动换行
    - 所有测试用例必须显式提供全局唯一 case_code
"""

from __future__ import annotations

import argparse
import json
import shutil
import sys
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment


# Excel 表头 -> cases.json 字段的映射
COLUMN_MAP = {
    "功能模块（一级模块）": "_module",
    "子功能（二级模块）": "_story_title",
    "测试项（三级模块）": "category",
    "自定义编号": "_case_id",
    "用例概要*": "summary",
    "优先级*": "priority",
    "前置条件": "precondition",
    "测试步骤": "_steps_text",
    "预期结果": "_expects_text",
}

WRAP_COLUMNS = {"测试步骤", "预期结果", "前置条件", "用例概要*"}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Fill test case template from cases.json")
    parser.add_argument("--template", required=True, help="Excel 模板路径")
    parser.add_argument("--input", required=True, help="cases.json 路径")
    parser.add_argument("--output", required=True, help="输出 xlsx 路径")
    return parser.parse_args()


def validate_case(case: dict, case_locator: str) -> None:
    """单条用例字段校验。"""
    required = ("category", "priority", "summary", "precondition", "steps", "expects")
    missing = [k for k in required if k not in case]
    if missing:
        sys.exit(f"[ERROR] {case_locator} 缺少字段: {', '.join(missing)}")

    if not isinstance(case["steps"], list) or not isinstance(case["expects"], list):
        sys.exit(f"[ERROR] {case_locator} steps/expects 必须为数组")

    if len(case["steps"]) == 0:
        sys.exit(f"[ERROR] {case_locator} steps 不能为空")

    if len(case["steps"]) != len(case["expects"]):
        sys.exit(
            f"[ERROR] {case_locator} steps({len(case['steps'])}) 与 expects({len(case['expects'])}) 长度不一致"
        )

    if case["priority"] not in ("P0", "P1", "P2"):
        sys.exit(f"[ERROR] {case_locator} priority 必须为 P0/P1/P2, 当前: {case['priority']}")


def format_numbered_lines(items: list[str]) -> str:
    return "\n".join(f"{i + 1}. {text}" for i, text in enumerate(items))


def flatten_cases(data: dict) -> list[dict]:
    """把 cases.json 摊平成可直接写入 Excel 的行字典列表。"""
    module = data.get("module")
    if not module:
        sys.exit("[ERROR] cases.json 缺少 module 字段")
    requirement_code = data.get("requirement_code")
    if not requirement_code:
        sys.exit("[ERROR] cases.json 缺少 requirement_code 字段")

    rows: list[dict] = []

    # 按故事处理
    for story in data.get("stories", []):
        story_code = story.get("story_code")
        story_title = story.get("story_title")
        if not story_code or not story_title:
            sys.exit("[ERROR] story 缺少 story_code 或 story_title")

        for idx, case in enumerate(story.get("cases", []), start=1):
            locator = f"{story_code} 第 {idx} 条"
            validate_case(case, locator)
            case_code = case.get("case_code")
            if not case_code:
                sys.exit(f"[ERROR] {locator} 缺少 case_code")
            rows.append(
                {
                    "_module": module,
                    "_story_title": story_title,
                    "_case_id": case_code,
                    "_steps_text": format_numbered_lines(case["steps"]),
                    "_expects_text": format_numbered_lines(case["expects"]),
                    **case,
                }
            )

    # 状态机用例（独立模块）
    sm = data.get("state_machine")
    if sm and sm.get("cases"):
        for idx, case in enumerate(sm["cases"], start=1):
            locator = f"状态机 第 {idx} 条"
            validate_case(case, locator)
            case_code = case.get("case_code")
            if not case_code:
                sys.exit(f"[ERROR] {locator} 缺少 case_code")
            rows.append(
                {
                    "_module": module,
                    "_story_title": "状态机验证",
                    "_case_id": case_code,
                    "_steps_text": format_numbered_lines(case["steps"]),
                    "_expects_text": format_numbered_lines(case["expects"]),
                    **case,
                }
            )

    if not rows:
        sys.exit("[ERROR] cases.json 中没有任何测试用例")

    return rows


def fill_workbook(template_path: Path, output_path: Path, rows: list[dict]) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy(template_path, output_path)

    wb = load_workbook(output_path)
    ws = wb.active

    headers = [cell.value for cell in ws[1]]
    header_to_col = {h: i + 1 for i, h in enumerate(headers) if h}

    missing_headers = [h for h in COLUMN_MAP if h not in header_to_col]
    if missing_headers:
        sys.exit(f"[ERROR] 模板缺少表头列: {', '.join(missing_headers)}")

    wrap_alignment = Alignment(wrap_text=True, vertical="top")

    for row_idx, row_data in enumerate(rows, start=2):
        for header, json_key in COLUMN_MAP.items():
            col = header_to_col[header]
            cell = ws.cell(row=row_idx, column=col, value=row_data.get(json_key, ""))
            if header in WRAP_COLUMNS:
                cell.alignment = wrap_alignment

    wb.save(output_path)


def main() -> None:
    args = parse_args()
    template_path = Path(args.template)
    input_path = Path(args.input)
    output_path = Path(args.output)

    if not template_path.exists():
        sys.exit(f"[ERROR] 模板不存在: {template_path}")
    if not input_path.exists():
        sys.exit(f"[ERROR] 输入 JSON 不存在: {input_path}")

    with input_path.open("r", encoding="utf-8") as f:
        data = json.load(f)

    rows = flatten_cases(data)
    fill_workbook(template_path, output_path, rows)

    story_counts: dict[str, int] = {}
    for r in rows:
        story_counts[r["_story_title"]] = story_counts.get(r["_story_title"], 0) + 1

    print(f"[OK] 已生成: {output_path}")
    print(f"[OK] 用例总数: {len(rows)}")
    print("[OK] 分布:")
    for title, count in story_counts.items():
        print(f"  - {title}: {count} 条")


if __name__ == "__main__":
    main()
